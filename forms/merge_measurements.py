from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook  # type: ignore
from rapidfuzz import fuzz, process  # type: ignore

from forms.extract_characteristics import Characteristic
from forms.mappings import CmmMapping, load_cmm_mapping


@dataclass
class Measurement:
    char_id: str
    description: str
    nominal: Optional[float]
    tolerance: Optional[float]
    unit: Optional[str]
    measured: Optional[float]
    instrument: Optional[str]
    pass_fail: Optional[bool]
    provenance: Dict[str, str]


def _read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def _read_xlsx(path: str, sheet: Optional[str] = None) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, str]] = []
    for ridx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        d: Dict[str, str] = {}
        for cidx, cell in enumerate(row):
            key = headers[cidx] if cidx < len(headers) else f"col{cidx+1}"
            val = cell.value
            d[key] = str(val) if val is not None else ""
        d["_row_number"] = str(ridx)
        rows.append(d)
    return rows


def _to_float(s: Optional[str]) -> Optional[float]:
    if s is None or s == "":
        return None
    try:
        return float(s)
    except Exception:
        try:
            return float(str(s).replace(",", "").strip())
        except Exception:
            return None


def merge_cmm_with_characteristics(
    characteristics: List[Characteristic],
    cmm_paths: List[str],
    mapping: Optional[CmmMapping] = None,
    fuzzy_threshold: int = 80,
) -> List[Measurement]:
    mapping = mapping or load_cmm_mapping()

    # Load all rows from provided CMM sources
    rows: List[Tuple[str, Dict[str, str]]] = []
    for path in cmm_paths:
        ext = os.path.splitext(path)[1].lower()
        if ext == ".csv":
            for r in _read_csv(path):
                r["_source_file"] = os.path.basename(path)
                rows.append((path, r))
        elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            for r in _read_xlsx(path):
                r["_source_file"] = os.path.basename(path)
                rows.append((path, r))
        else:
            continue

    # Index rows by explicit characteristic id when available
    id_index: Dict[str, Dict[str, str]] = {}
    for _path, row in rows:
        cid = (row.get(mapping.characteristic_id) or "").strip()
        if cid:
            id_index[cid.lower()] = row

    measurements: List[Measurement] = []
    # Prepare description candidates for fuzzy search
    row_descs = {i: (r.get(mapping.description or "Description", "")) for i, (_p, r) in enumerate(rows)}

    for ch in characteristics:
        matched_row: Optional[Dict[str, str]] = None
        # 1) Try id exact
        if ch.char_id and ch.char_id.lower() in id_index:
            matched_row = id_index[ch.char_id.lower()]
        else:
            # 2) Fuzzy match by description
            candidates = [(i, desc) for i, desc in row_descs.items() if desc]
            if ch.description and candidates:
                choice, score, index = process.extractOne(
                    ch.description, [c[1] for c in candidates], scorer=fuzz.token_sort_ratio
                )
                if score >= fuzzy_threshold:
                    matched_row = rows[candidates[index][0]][1]

        measured = _to_float(matched_row.get(mapping.measured)) if matched_row else None
        unit = matched_row.get(mapping.unit) if matched_row else (ch.unit or None)
        instrument = matched_row.get(mapping.instrument) if matched_row else None

        # pass/fail using nominal ± tolerance
        pf: Optional[bool] = None
        if ch.nominal is not None and ch.tolerance is not None and measured is not None:
            pf = (ch.nominal - ch.tolerance) <= measured <= (ch.nominal + ch.tolerance)

        provenance = {
            "source_file": matched_row.get("_source_file") if matched_row else "",
            "row_number": matched_row.get("_row_number") if matched_row else "",
        }

        measurements.append(
            Measurement(
                char_id=ch.char_id,
                description=ch.description,
                nominal=ch.nominal,
                tolerance=ch.tolerance,
                unit=unit,
                measured=measured,
                instrument=instrument,
                pass_fail=pf,
                provenance=provenance,
            )
        )

    return measurements
